import os
import wave
import  glob
import soundfile
ID = 3
tag = ['', 'dangbei', 'JmGO_H790', 'antiEpidemic002'][ID]
sname = ['20210620',][0]
dir = f'D:/data/ASR_DATA/weekly_data/{sname}/'
dir_wav  = dir + tag

def proc_pcm(sdir = dir_wav,
             tdir=dir + tag + '_' + sname,
             ptsv=f'{dir}/{tag}_{sname}.tsv'):
    if not os.path.exists(tdir):
        os.mkdir(tdir)
    names = glob.glob(sdir + '**/*.pcm', recursive=True)
    fw = open(ptsv, 'w', encoding='utf-8')
    total_duration = 0
    for i, name in enumerate(names):
        name = name.replace('\\', '/')
        with open(name, 'rb') as pcmfile:
            pcmdata = pcmfile.read()
        pos = name.rfind('/')
        pcm_path = os.path.join(tdir, name[pos+1:]).replace('.pcm', '.wav')
        if not os.path.exists(pcm_path):
            with wave.open(pcm_path , 'wb') as wavfile:
                wavfile.setparams((1, 2, 16000, 0, 'NONE', 'NONE'))
                wavfile.writeframes(pcmdata)
        frames = soundfile.info(pcm_path).frames
        if 10000 >= frames or frames >= 500000:
            print('skip.', name, frames)
            continue
        pcm_path = f'/audiodata6/zhaoang/workspace/azero_asr_ripple_test/{tag}_{sname}/' + os.path.basename(pcm_path)
        fw.write(f'{pcm_path}\t{frames}\t{16000}\n')
        total_duration += frames/ 16000
    print(total_duration / 3600)

import openpyxl
def generate_text_scp(stext=f'{dir}/{sname}.xlsx'):
    workbook = openpyxl.load_workbook(stext)
    print(workbook.sheetnames)
    sheet = workbook[workbook.sheetnames[ID]]
    tscript = f'{dir}/{tag}_{sname}.script'

    tsv = {}
    for line in open(f'{dir}/{tag}_{sname}.tsv', 'r',encoding='utf-8'):
        name, frames, sr= line.strip().split('\t')
        name = os.path.basename(name)
        tsv[name] = frames+'\t'+sr

    with open(tscript, 'w', encoding='utf-8') as fw_script:
        for i in range(1, len(sheet['C'])):
            name = f'/audiodata6/zhaoang/workspace/azero_asr_ripple_test/{tag}_{sname}/' + str(sheet['E'][i].value).replace('.pcm', '.wav')
            asr_text = str(sheet['G'][i].value)
            skill = None#sheet['G'][i].value if sheet['K'][i].value is None else sheet['K'][i].value #20210320 K

            base_name = os.path.basename(name)

            if base_name not in tsv:
                continue
            if skill is None:
                skill = 'default'
            #跳过 有unk 或者 tts的句子
            if asr_text.find('<') != -1:
                continue
            fw_script.write(name + '\t' + asr_text + '\t' + skill + '\t'+ tsv[base_name] +'\n')


proc_pcm()
generate_text_scp()
